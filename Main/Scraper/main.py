from bs4 import BeautifulSoup
import requests
import time
import pprint
from openpyxl import load_workbook
from openpyxl import workbook

def get_data():
    html_text = requests.get("https://app2.msci.com/eqb/custom_indexes/sg_performance.html").text

    soup = BeautifulSoup(html_text,"html.parser")

    idList = ["D_closing_index","D_constituents", "D_sectors"]
    
    #Open the excel file
    filename = "D:/SMU/Year 2 Term 2/Spreadsheet/Group project/Main/Scraper/SGSTOCKS.xlsx"
    wb = load_workbook(filename)
    ws = wb.active
    ws.title = "Data"
    rowNum = 1

    for pageid in idList :
        data = soup.find('div', id=pageid)

        #heading
        headings = data.find_all('b')
        for heading in headings:
            index = 0
            print(heading.text)
            #category
            category_Unprocessed = data.find_all('th')
            category = []
            for c in category_Unprocessed:
                category.append(c.text)

            #data
            sheet_data = []
            sheet_unprocessed = data.find_all('td')
            for s in sheet_unprocessed:
                sheet_data.append(s.text)

            #Insert heading
            c = ws.cell(row=rowNum, column=1)
            c.value = heading.text
            rowNum += 2 

            #Insert category
            if "Sector Weights" in heading.text:
                temp = 0
                while temp < 2:
                    c = ws.cell(row=rowNum, column=temp+1)
                    c.value = category[temp]
                    temp += 1
                rowNum += 1
            elif "Industry" in heading.text:
                temp = 2
                while temp < 4:
                    c = ws.cell(row=rowNum, column=temp-1)
                    c.value = category[temp]
                    temp += 1
                rowNum += 1
            else: 
                temp = 0
                while temp < len(category):
                    c = ws.cell(row=rowNum, column=temp+1)
                    c.value = category[temp]
                    temp += 1
                rowNum += 1
            
            #insert data 
            col_end = len(category)+1
            cell_col = 1

            if pageid == "D_closing_index":
                index_end = 6
            elif pageid == "D_constituents":
                index_end = 228
            else :
                if "Sector Weights" in heading.text:
                    col_end = 3
                    index_end = 14
                    index = 0
                else: 
                    col_end = 3
                    index_end = 32
                    index = 14

            while index < index_end:
                if cell_col == col_end:
                    cell_col = 1
                    rowNum += 1
                else:
                    c = ws.cell(row=rowNum, column=cell_col)
                    c.value = sheet_data[index]
                    index += 1
                    cell_col = cell_col+1

            rowNum += 3
        
    wb.save(filename)


if __name__ == "__main__":
    while True:
        get_data()
        time_wait = 24
        print("-"*100)
        print(f'Waiting {time_wait} hours')
        time.sleep(time_wait*60*60)






